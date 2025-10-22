import os
import time
import json
from pathlib import Path
from datetime import datetime
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework import status
from django.db import connection
from openpyxl import load_workbook
import water_logs.views as views
import xlrd
import logging

logger = logging.getLogger(__name__)

@csrf_exempt
def process_import_component_water_logs_data(request):
    try:
        body = json.loads(request.body.decode("utf-8"))
        company_id = request.auth.payload.get("company_id")
        user_id = body.get("user_id", 0)
        filename = body.get("filename")

        # Example hardcoded configs (replace with actual logic)
        liters_per_cane = 20  # e.g., 1 cane = 20 liters
        modified_on = int(time.time()* 1000);

        file_path = Path(settings.MEDIA_ROOT) / "upload_templates" / filename
        if not file_path.exists():
            return JsonResponse({ "status": False, "message": f"Error#01 File not found: {filename}", "data": {}}, status=status.HTTP_400_BAD_REQUEST);

        # --- Read Excel file ---
        try:
            if filename.endswith(".xlsx"):
                wb = load_workbook(file_path, read_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
            else:
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheet_by_index(0)
                rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        except Exception as e:
            return JsonResponse({"status": False,"message": f"Error#02 Unable to read Excel: {str(e)}","data": {}}, status=status.HTTP_500_INTERNAL_SERVER_ERROR);

        if not rows:
            return JsonResponse({"status": False,"message": "Error#03 No data found in file","data": {}}, status=status.HTTP_400_BAD_REQUEST)

        superadmin_response = views.get_superadmin_details({"company_id": company_id, "user_id": user_id});
        superadmin_details = superadmin_response['superadmin_data'][0] if superadmin_response and len(superadmin_response['superadmin_data']) > 0 else {}
        water_department = superadmin_details['water_department'] if superadmin_details and superadmin_details['water_department'] and superadmin_details['water_department'] == 1 else 0;

        expected_headers = ['User Name', 'Water Ids', 'User Ids', 'Month'];
        if water_department:
            expected_headers.append('Water Cane');
        else:
            expected_headers.append('Liter');

        actual_headers = [str(h).strip() for h in rows[0] if h]

        missing_headers = [h for h in expected_headers if h not in actual_headers]
        if missing_headers:
            return JsonResponse({"status": False,"message": f"Error#04 Missing columns: {', '.join(missing_headers)}","data": {}}, status=status.HTTP_400_BAD_REQUEST)

        # Index lookup
        water_cane_idx = 0;
        liter_idx = 0;
        water_id_idx = actual_headers.index("Water Ids");
        user_id_idx = actual_headers.index("User Ids");
        if water_department:
            water_cane_idx = actual_headers.index("Water Cane");
        else:
            liter_idx = actual_headers.index("Liter");

        updated_count = 0
        error_rows = []

        with connection.cursor() as cursor:
            user_response = views.get_user_details({"company_id": company_id, "user_id": user_id});
            user_details = user_response['user_data'][0] if user_response and len(user_response['user_data']) > 0 else {}
            rate_per_key = 'rate_per_cane' if superadmin_details['water_department'] and superadmin_details['water_department'] == 1 else 'rate_per_liter';
            column_key = 'water_cane' if superadmin_details['water_department'] and superadmin_details['water_department'] == 1 else 'liters';
            value_rate_per = user_details[rate_per_key] if user_details else 20 # default value of rate_per_cane

            for i, row in enumerate(rows[1:], start=2):
                try:
                    water_cane = 0;
                    liters = 0;
                    w_id = int(row[water_id_idx] or 0)
                    u_id = int(row[user_id_idx] or 0)
                    if water_department:
                        water_cane = float(row[water_cane_idx] or 0)
                    else:
                        liters = float(row[liter_idx] or 0)

                    if liters > 0 and superadmin_details and superadmin_details.get('water_department') == 0:
                        liters = float(liters)
                        water_cane = views.calculate_water_cane(liters, liters_per_cane)

                    elif water_cane > 0 and superadmin_details and superadmin_details.get('water_department') == 1:
                        water_cane = float(water_cane)
                        liters = views.calculate_water_liters(water_cane, liters_per_cane)

                    if u_id == user_id:
                        update_query = """
                            UPDATE muc_water_logs
                            SET liters = %s, water_cane = %s, modified_on = %s
                            WHERE company_id = %s AND user_id = %s AND water_id = %s
                        """
                        params = [liters, water_cane, modified_on, company_id, user_id, w_id]
                        cursor.execute(update_query, params)

                        if cursor.rowcount > 0:
                            updated_count += 1
                        else:
                            error_rows.append({
                                "row": i,
                                "reason": "No matching record found"
                            })
                    else:
                        error_rows.append({
                            "row": i,
                            "reason": f"User ID mismatch (Excel: {u_id}, Input: {user_id})"
                        })

                except Exception as e:
                    logger.error(f"Row {i} failed: {e}")
                    error_rows.append({"row": i, "reason": str(e)})

            return JsonResponse({ "status": True, "message": f"Data successfully updated.", "data": {"updated_count": updated_count},"errors": error_rows }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return JsonResponse({
            "status": False,
            "message": f"Unexpected error: {str(e)}",
            "data": {}
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
