import os
import time
from pathlib import Path
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
import xlrd

@csrf_exempt
def process_upload_excel_file(request):
    try:
        file = request.FILES.get('file')

        # --- Save file to media/upload_templates ---
        upload_root = Path(settings.MEDIA_ROOT) / 'upload_templates'
        upload_root.mkdir(parents=True, exist_ok=True)

        filename = f"{int(time.time())}_{file.name}"
        filepath = upload_root / filename

        with open(filepath, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        # --- Read Excel headers ---
        file_headers = []
        try:
            if filename.endswith('.xlsx'):
                wb = load_workbook(filepath, read_only=True)
                sheet = wb.active
                first_row = next(sheet.iter_rows(values_only=True))
                file_headers = [cell for cell in first_row if cell is not None]
            else:
                wb = xlrd.open_workbook(filepath)
                sheet = wb.sheet_by_index(0)
                file_headers = sheet.row_values(0)
        except Exception as e:
            return JsonResponse({'status': False, 'message': f'Error reading Excel: {str(e)}'}, status=500)

        # --- Build response ---
        return {
            'status': True,
            'fileHeaders': file_headers,
            'filename': filename,
            'filepath': str(filepath),
            'relative_path': str(filepath.relative_to(settings.MEDIA_ROOT))
        }

    except Exception as e:
        return {'status': False, 'message': f'Unexpected error: {str(e)}'}
